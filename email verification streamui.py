import re
import dns.resolver
import smtplib
import socket
import csv
import os
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# Color definitions
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
# Initialize workbook
wb = Workbook()
ws = wb.active
ws.title = "Validation Results"
ws.append(["Email", "Message", "Status"])

def is_valid_email_format(email):
    regex = r'^[\w\.-]+@([\w\.-]+\.\w+)$'
    return re.match(regex, email)

def check_dns_records(domain):
    records = {'A': None, 'NS': None, 'CNAME': None}
    try:
        a = dns.resolver.resolve(domain, 'A')
        records['A'] = [r.address for r in a]
    except: pass
    try:
        ns = dns.resolver.resolve(domain, 'NS')
        records['NS'] = [r.to_text() for r in ns]
    except: pass
    try:
        cname = dns.resolver.resolve(domain, 'CNAME')
        records['CNAME'] = [r.to_text() for r in cname]
    except: pass
    return records

def check_mx_records(domain):
    try:
        mx = dns.resolver.resolve(domain, 'MX')
        records = [
            (r.preference, str(r.exchange).rstrip('.'))
            for r in mx
            if 'localhost' not in str(r.exchange) and '127.' not in str(r.exchange)
        ]
        return sorted(records, key=lambda x: x[0])
    except:
        return None

def smtp_probe(email, mail_server, debug=False):
    try:
        if debug:
            print(f"🔌 Connecting to {mail_server} ...")
        server = smtplib.SMTP(mail_server, 25, timeout=10)
        if debug:
            server.set_debuglevel(1)
        server.ehlo()
        server.mail('check3@gmail.com')
        code, message = server.rcpt(email)
        server.quit()

        msg_text = message.decode() if isinstance(message, bytes) else str(message)

        if debug:
            print(f"📨 RCPT TO returned: {code} - {msg_text}")

        if code == 250:
            return True, "✅ Deliverable"
        elif code == 550:
            return False, "❌ Mailbox does not exist"
        elif code in [553, 554]:
            return None, f"⚠️ Blocked or IP reputation issue: {msg_text}"
        else:
            return None, f"⚠️ Unhandled response: {msg_text}"

    except Exception as e:
        return None, f"❌ SMTP connection error: {e}"

def validate_email(email, check_smtp=True, debug=False):
    if not is_valid_email_format(email):
        return False, "❌ Invalid email format"

    domain = email.split('@')[1]
    dns_records = check_dns_records(domain)
    if not any(dns_records.values()):
        return False, "❌ Domain does not exist or has no DNS records"

    mx_records = check_mx_records(domain)
    if not mx_records:
        return False, "❌ No MX records found (cannot receive emails)"

    if debug:
        print("🔍 MX records (by priority):")
        for priority, host in mx_records:
            try:
                ip = socket.gethostbyname(host)
                print(f"{host} (priority {priority}) → {ip}")
            except:
                print(f"{host} (priority {priority}) → ❌ could not resolve")

    if check_smtp:
        rejected = False
        for priority, host in mx_records:
            if debug:
                print(f"🔌 Probing {host} (priority {priority}) ...")
            result, message = smtp_probe(email, host, debug=debug)
            if result is True:
                return True, f"✅ Deliverable via {host}"
            elif result is False:
                rejected = True
                if debug:
                    print(message)
            else:
                if debug:
                    print(message)
        if rejected:
            return False, "❌ Mailbox rejected by all MX servers"
        return None, "⚠️ Email might be invalid (SMTP inconclusive)"
    
    return True, "✅ Format, domain, and MX valid (SMTP not checked)"

def get_status_and_score(valid, message):
    if valid is True:
        return "Valid", 1.0
    elif valid is False:
        return "Invalid", 0.0
    else:
        return "Likely Invalid", 0.5  # For None/inconclusive

# === Main Logic ===
def main():
    st.title("📧 Email Validator")

    st.markdown("Enter a single email or upload a file containing a list of emails.")

    check_smtp = st.checkbox("Use SMTP Check (slower, more accurate)?", value=True)

    input_method = st.radio("Select Input Method", ("Single Email", "Upload File"))

    if input_method == "Single Email":
        email = st.text_input("Enter Email Address")
        if st.button("Validate"):
            if email:
                valid, msg = validate_email(email, check_smtp=check_smtp)
                status, _ = get_status_and_score(valid, msg)
                color = {"Valid": "green", "Invalid": "red", "Likely Invalid": "orange"}[status]
                st.markdown(f"**Result:** <span style='color:{color}'>{status}</span> - {msg}", unsafe_allow_html=True)
            else:
                st.warning("Please enter an email address.")

    else:
        uploaded_file = st.file_uploader("Upload Excel or CSV file", type=["csv", "xlsx", "xls"])
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith(".csv"):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file)

                if "Email address" not in df.columns:
                    st.error("Uploaded file must contain 'Email address' column.")
                    return

                st.success(f"Found {len(df)} email(s). Starting validation...")

                statuses, messages = [], []
                for email in df["Email address"]:
                    email = str(email).strip()
                    valid, msg = validate_email(email, check_smtp=check_smtp)
                    status, _ = get_status_and_score(valid, msg)
                    statuses.append(status)
                    messages.append(msg)

                df["Validation Status"] = statuses
                df["Validation Message"] = messages

                # Save to Excel with color formatting
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Validation Results"
                ws.append(df.columns.tolist())

                for idx, row in df.iterrows():
                    row_values = row.tolist()
                    ws.append(row_values)
                    status = row["Validation Status"]
                    fill = green_fill if status == "Valid" else red_fill if status == "Invalid" else yellow_fill
                    ws.cell(row=idx+2, column=df.columns.get_loc("Validation Status")+1).fill = fill

                wb.save(output)
                output.seek(0)

                st.success("Validation completed!")
                st.download_button("📥 Download Validated Results", data=output, file_name="validated_emails.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Something went wrong: {e}")


if __name__ == "__main__":
    main()
