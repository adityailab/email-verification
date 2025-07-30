import re
import dns.resolver
import smtplib
import socket
import csv
import os
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# Color definitions
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
# Initialize workbook
wb = Workbook()
ws = wb.active
ws.title = "Validation Results"
ws.append(["Email", "Message", "Status"])

def is_valid_email_format(email):
    regex = r'^[\w\.-]+@([\w\.-]+\.\w+)$'
    return re.match(regex, email)

def check_dns_records(domain):
    records = {'A': None, 'NS': None, 'CNAME': None}
    try:
        a = dns.resolver.resolve(domain, 'A')
        records['A'] = [r.address for r in a]
    except: pass
    try:
        ns = dns.resolver.resolve(domain, 'NS')
        records['NS'] = [r.to_text() for r in ns]
    except: pass
    try:
        cname = dns.resolver.resolve(domain, 'CNAME')
        records['CNAME'] = [r.to_text() for r in cname]
    except: pass
    return records

def check_mx_records(domain):
    try:
        mx = dns.resolver.resolve(domain, 'MX')
        records = [
            (r.preference, str(r.exchange).rstrip('.'))
            for r in mx
            if 'localhost' not in str(r.exchange) and '127.' not in str(r.exchange)
        ]
        return sorted(records, key=lambda x: x[0])
    except:
        return None

def smtp_probe(email, mail_server, debug=False):
    try:
        if debug:
            print(f"🔌 Connecting to {mail_server} ...")
        server = smtplib.SMTP(mail_server, 25, timeout=10)
        if debug:
            server.set_debuglevel(1)
        server.ehlo()
        server.mail('check3@gmail.com')
        code, message = server.rcpt(email)
        server.quit()

        msg_text = message.decode() if isinstance(message, bytes) else str(message)

        if debug:
            print(f"📨 RCPT TO returned: {code} - {msg_text}")

        if code == 250:
            return True, "✅ Deliverable"
        elif code == 550:
            return False, "❌ Mailbox does not exist"
        elif code in [553, 554]:
            return None, f"⚠️ Blocked or IP reputation issue: {msg_text}"
        else:
            return None, f"⚠️ Unhandled response: {msg_text}"

    except Exception as e:
        return None, f"❌ SMTP connection error: {e}"

def validate_email(email, check_smtp=True, debug=False):
    if not is_valid_email_format(email):
        return False, "❌ Invalid email format"

    domain = email.split('@')[1]
    dns_records = check_dns_records(domain)
    if not any(dns_records.values()):
        return False, "❌ Domain does not exist or has no DNS records"

    mx_records = check_mx_records(domain)
    if not mx_records:
        return False, "❌ No MX records found (cannot receive emails)"

    if debug:
        print("🔍 MX records (by priority):")
        for priority, host in mx_records:
            try:
                ip = socket.gethostbyname(host)
                print(f"{host} (priority {priority}) → {ip}")
            except:
                print(f"{host} (priority {priority}) → ❌ could not resolve")

    if check_smtp:
        rejected = False
        for priority, host in mx_records:
            if debug:
                print(f"🔌 Probing {host} (priority {priority}) ...")
            result, message = smtp_probe(email, host, debug=debug)
            if result is True:
                return True, f"✅ Deliverable via {host}"
            elif result is False:
                rejected = True
                if debug:
                    print(message)
            else:
                if debug:
                    print(message)
        if rejected:
            return False, "❌ Mailbox rejected by all MX servers"
        return None, "⚠️ Email might be invalid (SMTP inconclusive)"
    
    return True, "✅ Format, domain, and MX valid (SMTP not checked)"

def get_status_and_score(valid, message):
    if valid is True:
        return "Valid", 1.0
    elif valid is False:
        return "Invalid", 0.0
    else:
        return "Likely Invalid", 0.5  # For None/inconclusive

# === Main Logic ===
def main():
    path_or_email = input("Enter an email address or path to Excel/CSV file: ").strip()
    check_smtp = input("Use SMTP check? (y/n): ").strip().lower() == 'y'

    if is_valid_email_format(path_or_email):
        # Single email
        status, message = validate_email(path_or_email, check_smtp=check_smtp)
        print(f"\n📧 {path_or_email} → {status}: {message}")
        return

    if not os.path.isfile(path_or_email):
        print("❌ Not a valid email or file.")
        return

    # === Load file ===
    if path_or_email.endswith(".csv"):
        df = pd.read_csv(path_or_email)
    elif path_or_email.endswith((".xlsx", ".xls")):
        df = pd.read_excel(path_or_email)
    else:
        print("❌ Unsupported file format.")
        return

    if "Email address" not in df.columns:
        print("❌ Missing 'Email address' column.")
        return

    print(f"\n📄 Validating {len(df)} emails...\n")

    statuses, messages = [], []
    for email in df["Email address"]:
        email = str(email).strip()
        # status, msg = validate_email(email, check_smtp=check_smtp)
        valid, msg = validate_email(email, check_smtp=check_smtp)
        status = "Valid" if valid is True else "Invalid" if valid is False else "Likely Invalid"

        print(f"{email} → {status}: {msg}")
        statuses.append(status)
        messages.append(msg)

    df["Validation Status"] = statuses
    df["Validation Message"] = messages

    output_file = "validated_results.xlsx"
    df.to_excel(output_file, index=False)

    # === Apply Color Coding ===
    wb = load_workbook(output_file)
    ws = wb.active
    status_col_letter = chr(ord('A') + df.columns.get_loc("Validation Status"))

    for row in range(2, ws.max_row + 1):
        status_cell = f"{status_col_letter}{row}"
        status_value = ws[status_cell].value
        if status_value == "Valid":
            ws[status_cell].fill = green_fill
        elif status_value == "Invalid":
            ws[status_cell].fill = red_fill
        else:
            ws[status_cell].fill = yellow_fill

    wb.save(output_file)
    print(f"\n✅ Results saved to: {output_file}")

if __name__ == "__main__":
    main()
